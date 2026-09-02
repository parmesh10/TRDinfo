from django.contrib import admin, messages
from django.db import transaction
from django.http import HttpResponseRedirect
from django.shortcuts import render
from django.urls import path

from openpyxl import load_workbook

from .models import (
    Question,
    Exam,
    ExamAttempt,
    ExamAnswer,
)


@admin.register(Question)
class QuestionAdmin(admin.ModelAdmin):

    list_display = (
        "id",
        "question_text",
        "difficulty",
        "correct_answer",
        "active",
        "created_at",
    )

    list_filter = (
        "difficulty",
        "active",
    )

    search_fields = (
        "question_text",
        "option_a",
        "option_b",
        "option_c",
        "option_d",
    )

    def get_urls(self):
        urls = super().get_urls()

        custom_urls = [
            path(
                "upload-questions/",
                self.admin_site.admin_view(
                    self.upload_questions
                ),
                name="exam_question_upload",
            ),
        ]

        return custom_urls + urls
        
    
        
    def changelist_view(self, request, extra_context=None):
        from django.urls import reverse

        extra_context = extra_context or {}

        extra_context["upload_questions_url"] = reverse(
            "admin:exam_question_upload"
        )

        return super().changelist_view(
            request,
            extra_context=extra_context
        )

    def upload_questions(self, request):

        if request.method == "POST":

            excel_file = request.FILES.get("excel_file")

            if not excel_file:
                messages.error(
                    request,
                    "Please select an Excel file."
                )
                return HttpResponseRedirect(request.path)

            if not excel_file.name.lower().endswith(".xlsx"):
                messages.error(
                    request,
                    "Only .xlsx files are supported."
                )
                return HttpResponseRedirect(request.path)

            try:
                workbook = load_workbook(
                    excel_file,
                    read_only=True,
                    data_only=True
                )

                worksheet = workbook.active

                rows = list(
                    worksheet.iter_rows(
                        values_only=True
                    )
                )

                if not rows:
                    messages.error(
                        request,
                        "The Excel file is empty."
                    )
                    return HttpResponseRedirect(request.path)

                headers = [
                    str(cell).strip().lower()
                    if cell is not None
                    else ""
                    for cell in rows[0]
                ]

                required_headers = [
                    "question",
                    "option a",
                    "option b",
                    "option c",
                    "option d",
                    "correct answer",
                ]

                missing = [
                    h for h in required_headers
                    if h not in headers
                ]

                if missing:
                    messages.error(
                        request,
                        "Missing columns: "
                        + ", ".join(missing)
                    )
                    return HttpResponseRedirect(request.path)

                column_index = {
                    h: headers.index(h)
                    for h in headers
                }

                imported = 0
                skipped = 0
                errors = []

                for row_number, row in enumerate(
                    rows[1:],
                    start=2
                ):

                    def value(column):
                        index = column_index.get(column)

                        if index is None:
                            return ""

                        if index >= len(row):
                            return ""

                        if row[index] is None:
                            return ""

                        return str(row[index]).strip()

                    question_text = value("question")
                    option_a = value("option a")
                    option_b = value("option b")
                    option_c = value("option c")
                    option_d = value("option d")
                    correct_answer = value(
                        "correct answer"
                    ).upper()

                    explanation = value("explanation")

                    difficulty = value(
                        "difficulty"
                    ).upper() or "MEDIUM"

                    active_value = value(
                        "active"
                    ).lower()

                    if not question_text:
                        errors.append(
                            f"Row {row_number}: "
                            "Question is empty."
                        )
                        continue

                    if not option_a or not option_b:
                        errors.append(
                            f"Row {row_number}: "
                            "Option A/B is empty."
                        )
                        continue

                    if not option_c or not option_d:
                        errors.append(
                            f"Row {row_number}: "
                            "Option C/D is empty."
                        )
                        continue

                    if correct_answer not in (
                        "A",
                        "B",
                        "C",
                        "D",
                    ):
                        errors.append(
                            f"Row {row_number}: "
                            "Correct Answer must be A, B, C or D."
                        )
                        continue

                    if difficulty not in (
                        "EASY",
                        "MEDIUM",
                        "HARD",
                    ):
                        errors.append(
                            f"Row {row_number}: "
                            "Difficulty must be EASY, MEDIUM or HARD."
                        )
                        continue

                    active = active_value not in (
                        "no",
                        "false",
                        "0",
                        "inactive",
                    )

                    if Question.objects.filter(
                        question_text__iexact=question_text
                    ).exists():
                        skipped += 1
                        continue

                    Question.objects.create(
                        question_text=question_text,
                        option_a=option_a,
                        option_b=option_b,
                        option_c=option_c,
                        option_d=option_d,
                        correct_answer=correct_answer,
                        explanation=explanation,
                        difficulty=difficulty,
                        active=active,
                    )

                    imported += 1

                workbook.close()

                messages.success(
                    request,
                    f"Import completed: "
                    f"{imported} imported, "
                    f"{skipped} duplicates skipped, "
                    f"{len(errors)} errors."
                )

                for error in errors[:20]:
                    messages.warning(request, error)

                return HttpResponseRedirect(request.path)

            except Exception as e:

                messages.error(
                    request,
                    f"Error reading Excel: {e}"
                )

                return HttpResponseRedirect(request.path)

        context = dict(
            self.admin_site.each_context(request),
            title="Upload Questions",
            opts=self.model._meta,
        )

        return render(
            request,
            "admin/exam/question/upload_questions.html",
            context,
        )

@admin.register(Exam)
class ExamAdmin(admin.ModelAdmin):

    list_display = (
        "name",
        "number_of_questions",
        "duration_minutes",
        "pass_percentage",
        "negative_marking",
        "active",
    )

    list_filter = (
        "active",
        "negative_marking",
    )


@admin.register(ExamAttempt)
class ExamAttemptAdmin(admin.ModelAdmin):

    list_display = (
        "candidate_name",
        "employee_number",
        "exam",
        "total_questions",
        "correct_answers",
        "wrong_answers",
        "unanswered_questions",
        "score",
        "percentage",
        "passed",
        "started_at",
        "submitted_at",
    )

    list_filter = (
        "exam",
        "passed",
    )

    search_fields = (
        "candidate_name",
        "employee_number",
    )


@admin.register(ExamAnswer)
class ExamAnswerAdmin(admin.ModelAdmin):

    list_display = (
        "attempt",
        "question",
        "selected_answer",
        "is_correct",
        "marks_obtained",
    )

    list_filter = (
        "is_correct",
    )